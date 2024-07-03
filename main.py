# coding:utf-8
import os
import sys
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from PyQt6 import QtCore, QtWidgets
from PyQt6.QtCore import Qt, QUrl, QEasingCurve
from PyQt6.QtGui import QIcon, QDesktopServices
from PyQt6.QtWidgets import QApplication, QFrame, QHBoxLayout, QVBoxLayout, QBoxLayout, QTableWidgetItem, QSizePolicy
from qfluentwidgets import (NavigationItemPosition, MessageBox, setTheme, Theme, FluentWindow,
                            NavigationAvatarWidget, qrouter, SubtitleLabel, setFont, InfoBadge,
                            InfoBadgePosition, FluentBackgroundTheme, SwitchSettingCard, SettingCard, PrimaryPushButton,
                            FluentIcon, SplitPushButton, TransparentTogglePushButton, BodyLabel, TransparentPushButton,
                            LineEdit, SingleDirectionScrollArea, TableWidget, RoundMenu, Action, PushButton, Flyout,
                            InfoBarIcon, FlyoutAnimationType, FlyoutView, CardWidget)
from qfluentwidgets import FluentIcon as FIF
from qfluentwidgets import CheckBox, ComboBox, NavigationBar

def path_file(file):
    bundle_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
    path_to_file = os.path.join(bundle_dir, file)

    return path_to_file
class Widget(QFrame):

    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = SubtitleLabel(text, self)
        self.hBoxLayout = QHBoxLayout(self)

        setFont(self.label, 24)
        self.label.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.hBoxLayout.addWidget(self.label, 1, Qt.AlignmentFlag.AlignCenter)
        self.setObjectName(text.replace(' ', '-'))

        self.CheckBox = CheckBox(parent=parent)
        self.CheckBox.setGeometry(QtCore.QRect(330, 390, 92, 22))
        self.CheckBox.setObjectName("CheckBox")
        self.hBoxLayout.addWidget(self.CheckBox, 1, Qt.AlignmentFlag.AlignCenter)
class Widget2(QFrame):

    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = SubtitleLabel(text, self)
        self.hBoxLayout = QVBoxLayout(self)
        self.hBoxLayout.setContentsMargins(5,0,5,0)
        self.hBoxLayout.setDirection(QBoxLayout.Direction.TopToBottom)
        self.hBoxLayout.setSpacing(5)

        setFont(self.label, 30)
        self.label.setFixedSize(200, 50)
        self.label.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        self.hBoxLayout.addWidget(self.label, 1, )
        self.setObjectName(text.replace(' ', '-'))
        # self.hBoxLayout.setContentsMargins(0, 0, 0, 900)
class Basic_salary():
    def __init__(self):
        self.layout=QHBoxLayout()
        self.sshftext=TransparentPushButton(QIcon(path_file("icons/银行卡_bank-card-two.svg")),'社保+公积金')
        self.sshf=LineEdit()
        self.sshf.setPlaceholderText("填入社保+公积金的金额")
        self.sshf.setText("733")

        self.bstext = TransparentPushButton(QIcon(path_file("icons/金融_finance.svg")), '底薪')
        self.bs = LineEdit()
        self.bs.setPlaceholderText("填入底薪")
        self.bs.setText("4000")

        self.compulsory_class_hours = TransparentPushButton(QIcon(path_file("icons/时间_time.svg")), '义务课时')
        self.compulsory_class_hours = LineEdit()
        self.compulsory_class_hours.setPlaceholderText("填入义务课时")
        self.compulsory_class_hours.setText("18")

        self.layout.addWidget(self.sshftext,0,Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.sshf, 0, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.bstext, 0, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.bs, 0, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.compulsory_class_hours, 0, Qt.AlignmentFlag.AlignLeft)
class Classlist():
    def __init__(self):
        self.list = []
        self.set = SettingCard(icon=FluentIcon.PEOPLE, title='上课人数')
        self.set.hBoxLayout.setContentsMargins(5,0,5,0)
        self.numberofpeople = ComboBox()
        peopleitems = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']
        self.numberofpeople.addItems(peopleitems)

        self.frequency = ComboBox()
        houritems=[]
        for i in range(1, 100):
            houritems.append(str(i))
        self.frequency.addItems(houritems)

        self.junior=ComboBox()
        junioritems=['小学','7年级','8年级','9年级']
        self.junior.addItems(junioritems)

        self.set.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)
        self.set.hBoxLayout.addWidget(self.numberofpeople, 5, Qt.AlignmentFlag.AlignTop)
        self.set.hBoxLayout.addWidget(TransparentPushButton(QIcon(path_file("icons/时间_time.svg")),'上课次数'), 5, Qt.AlignmentFlag.AlignRight)
        self.set.hBoxLayout.addWidget(self.frequency, 5, Qt.AlignmentFlag.AlignTop)
        self.set.hBoxLayout.addWidget(self.junior, 5, Qt.AlignmentFlag.AlignTop)

class Salary_calculator(Widget2):
    def __init__(self, text: str, parent=None):
        super().__init__(text, parent)

        self.hBoxLayout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.card1=CardWidget()
        self.card1.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)

        self.card1layout = QVBoxLayout(self.card1)


        self.Teacher_Qualification_Certificate = SwitchSettingCard(icon=QIcon(path_file("icons/证书_certificate.svg")),title='教师资格证',)
        self.Teacher_Qualification_Certificate.setChecked(True)
        self.Teacher_Qualification_Certificate.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)
        self.card1layout.addWidget(self.Teacher_Qualification_Certificate, 1, )

        # self.hBoxLayout.addWidget(self.Teacher_Qualification_Certificate, 1, )

        self.Full_attendance = SwitchSettingCard(icon=QIcon(path_file("icons/日历_calendar.svg")), title='全勤')
        self.Full_attendance.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)
        self.Full_attendance.setChecked(True)
        self.card1layout.addWidget(self.Full_attendance, 1, )
        # self.hBoxLayout.addWidget(self.Full_attendance, 1,)

        self.summer_class = SwitchSettingCard(icon=QIcon(path_file("icons/火热_fire.svg")), title='暑假班')
        self.summer_class.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)
        self.summer_class.setChecked(True)
        self.card1layout.addWidget(self.summer_class, 1, )
        # self.hBoxLayout.addWidget(self.summer_class, 1, )
        self.hBoxLayout.addWidget(self.card1, 1, )

        self.card2 = CardWidget()
        self.card2.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        self.card2layout = QHBoxLayout(self.card2)

        self.action_buttons= QHBoxLayout()
        self.count = PrimaryPushButton(text='计算结果', icon=FluentIcon.UPDATE)
        self.count.setMaximumSize(QtCore.QSize(16777215, 35))
        self.count.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        self.count.clicked.connect(self.count_wages)

        # self.action_buttons.addWidget(self.count, 1,)
        self.card2layout.addWidget(self.count, 1, )

        self.check = SplitPushButton(text='历史记录', icon=FluentIcon.HISTORY)
        self.check.setMaximumSize(QtCore.QSize(16777215, 35))
        self.check.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        # self.action_buttons.addWidget(self.check, 1, Qt.AlignmentFlag.AlignLeft)
        self.card2layout.addWidget(self.check, 1, Qt.AlignmentFlag.AlignLeft)

        self.addlist=PrimaryPushButton(text='添加一条上课记录',icon=FluentIcon.ADD)
        self.addlist.setMaximumSize(QtCore.QSize(16777215, 35))
        self.addlist.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        self.addlist.clicked.connect(self.add_new_list)
        # self.action_buttons.addWidget(self.addlist, 1,)
        self.card2layout.addWidget(self.addlist, 1, )

        self.basicsalary=Basic_salary()
        self.hBoxLayout.addLayout(self.basicsalary.layout)

        # self.hBoxLayout.addLayout(self.action_buttons)
        self.hBoxLayout.addWidget(self.card2, 1, )
        self.hBoxLayout.setContentsMargins(0, 0, 0, 0)
        # self.hBoxLayout.setStretch(-1, 7)
        self.classlist= []
        self.listcount = 0
        self.history=RoundMenu(parent=self.check)
        # bundle_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
        # path_to_file = os.path.join(bundle_dir, 'salary_info.xlsx')
        wb = load_workbook(path_file('salary_info.xlsx')).active
        for row in wb.iter_rows(min_row=2):
            # 将每一行的数据合并为一个字符串
            row_data = [str(cell.value) for cell in row]
            row_string = ' '.join(row_data)
            self.history.addAction(Action(FluentIcon.CLOUD, row_string))
        self.check.setFlyout(self.history)
    def count_wages(self):
        sum=float(0)
        yiwu=float(self.basicsalary.compulsory_class_hours.text())
        payroll=[{'number': 1, 'money': 40,'grade':6}, {'number': 2, 'money': 45,'grade':6}, {'number': 3, 'money': 60,'grade':6}, ]

        sum=sum+int(self.basicsalary.bs.text())-int(self.basicsalary.sshf.text())
        if self.Teacher_Qualification_Certificate.isChecked():
            sum+=50
        if self.Full_attendance.isChecked():
            sum+=300
        calculationlist=[]
        for i in range(0,len(self.classlist)):
            calculationlist.append({'people':float(self.classlist[i].numberofpeople.currentText()), 'hours':float(self.classlist[i].frequency.currentText())*1.5})
            if self.classlist[i].junior.currentText()=='小学':
                calculationlist[i]['grade']=6
            if self.classlist[i].junior.currentText()=='7年级':
                calculationlist[i]['grade']=7
            if self.classlist[i].junior.currentText()=='8年级':
                calculationlist[i]['grade']=8
            if self.classlist[i].junior.currentText()=='9年级':
                calculationlist[i]['grade']=9
        calculationlist = sorted(calculationlist, key=lambda item: (item['people'], item['grade']))
        if not self.summer_class.isChecked():
            for i in range(0,len(calculationlist)):
                if calculationlist[i]['people']>2:
                    calculationlist[i]['hours'] = calculationlist[i]['hours'] / 1.5 * 2
        for i in range(0,len(calculationlist)):
            money=0
            if calculationlist[i]['people']<=2:
                money=payroll[int(calculationlist[i]['people']-1)]['money']
                if 6<calculationlist[i]['grade']<9:
                    money+=5
                elif calculationlist[i]['grade']==9:
                    money+=10

            if calculationlist[i]['people'] > 2:
                if calculationlist[i]['grade'] > 6:
                    money = payroll[2]['money'] + 5 * (calculationlist[i]['people'] - 3)+20
                else:
                    money = payroll[2]['money'] + 5 * (calculationlist[i]['people'] - 3)
            temp = {'people': calculationlist[i]['people'], 'hours': calculationlist[i]['hours'], 'money': money}
            calculationlist[i] = temp
        calculationlist = sorted(calculationlist, key=lambda item: (item['people'] != 1, item['money']))
        for i in range(0,len(calculationlist)):
            yiwu=yiwu-calculationlist[i]['hours']
            if yiwu >=0:
                continue
            if yiwu<0:
                sum+=calculationlist[i]['money']*(calculationlist[i]['hours'])
                continue

        print("总工资",sum)
        result=MessageBox('工资计算结果',f'总工资为{sum}元',self)
        result.yesButton.setText('耶！')
        result.yesButton.clicked.connect(self.whoknow)
        result.cancelButton.setText('好咯')
        result.show()

        wb = load_workbook(path_file("salary_info.xlsx"))
        ws = wb.active
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([current_time, sum])
        wb.save(path_file("salary_info.xlsx"))

        temp = RoundMenu(parent=self.check)
        wb = load_workbook(path_file("salary_info.xlsx")).active
        for row in wb.iter_rows(min_row=2):
            # 将每一行的数据合并为一个字符串
            row_data = [str(cell.value) for cell in row]
            row_string = ' '.join(row_data)
            temp.addAction(Action(FluentIcon.CLOUD, row_string))
        self.check.setFlyout(temp)



        print(calculationlist)

    def whoknow(self):
        result = MessageBox('也不知道谁弄的', '', self)
        result.yesButton.setText('厉害咯')
        result.cancelButton.setText('爱咯')
        result.show()
    def add_new_list(self):
        print('add_new_list')
        temp=Classlist()
        self.classlist.append(temp)
        # sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Expanding)
        #
        # self.classlist[self.listcount].set.SizePolicy(sizePolicy)
        self.hBoxLayout.addWidget(self.classlist[self.listcount].set, 0, )
        self.listcount += 1
        # self.hBoxLayout.setStretch(-1, 7)


class Window(FluentWindow):

    def __init__(self):
        super().__init__()

        # create sub interface
        self.homeInterface = Salary_calculator('工资计算器', self)
        # self.musicInterface = Widget('Music Interface', self)
        # self.videoInterface = Widget('Video Interface', self)
        # self.folderInterface = Widget('Folder Interface', self)
        # self.settingInterface = Widget('Setting Interface', self)
        # self.albumInterface = Widget('Album Interface', self)
        # self.albumInterface1 = Widget('Album Interface 1', self)
        # self.albumInterface2 = Widget('Album Interface 2', self)
        # self.albumInterface1_1 = Widget('Album Interface 1-1', self)


        self.initNavigation()
        self.initWindow()

    def initNavigation(self):
        self.addSubInterface(self.homeInterface, FIF.HOME, '工资计算器')
        # self.addSubInterface(self.musicInterface, FIF.MUSIC, 'Music library')
        # self.addSubInterface(self.videoInterface, FIF.VIDEO, 'Video library')
        #
        # self.navigationInterface.addSeparator()
        #
        # self.addSubInterface(self.albumInterface, FIF.ALBUM, 'Albums', NavigationItemPosition.SCROLL)
        # self.addSubInterface(self.albumInterface1, FIF.ALBUM, 'Album 1', parent=self.albumInterface)
        # self.addSubInterface(self.albumInterface1_1, FIF.ALBUM, 'Album 1.1', parent=self.albumInterface1)
        # self.addSubInterface(self.albumInterface2, FIF.ALBUM, 'Album 2', parent=self.albumInterface)
        # self.addSubInterface(self.folderInterface, FIF.FOLDER, 'Folder library', NavigationItemPosition.SCROLL)

        # add custom widget to bottom
        self.navigationInterface.addWidget(
            routeKey='avatar',
            widget=NavigationAvatarWidget('Aldly', path_file('icons/Aldly.png')),
            onClick=self.showMessageBox,
            position=NavigationItemPosition.BOTTOM,
        )

        # self.addSubInterface(self.settingInterface, FIF.SETTING, 'Settings', NavigationItemPosition.BOTTOM)


    def initWindow(self):
        self.resize(845, 714)
        self.setWindowIcon(QIcon(path_file('icons/Aldly.png')))
        self.setWindowTitle('乖宝工具箱')

        desktop = QApplication.screens()[0].availableGeometry()
        w, h = desktop.width(), desktop.height()
        self.move(w//2 - self.width()//2, h//2 - self.height()//2)
        self.setCustomBackgroundColor(*FluentBackgroundTheme.DEFAULT_BLUE)

    def showMessageBox(self):
        w = MessageBox(
            '支持作者🥰',
            '个人开发不易，如果这个项目帮助到了您，可以考虑请作者喝一瓶快乐水🥤。您的支持就是作者开发和维护项目的动力🚀',
            self
        )
        w.yesButton.setText('来啦老弟')
        w.cancelButton.setText('下次一定')
        w.yesButton.clicked.connect(lambda: self.showFlyout(w))
        w.show()

    def showFlyout(self,to):
        print('show flyout')
        view = FlyoutView(
            title='谢谢你的捐赠！',
            content="你要给我买快乐水吗？",
            image=path_file('icons/donation.jpg'),
            isClosable=False
        )

        # add button to view
        button = PushButton('关闭')
        button.setFixedWidth(120)

        view.addWidget(button, align=Qt.AlignmentFlag.AlignBottom)

        # adjust layout (optional)
        view.widgetLayout.insertSpacing(1, 5)
        view.widgetLayout.addSpacing(5)

        # show view
        w = Flyout.make(view, to.yesButton, self)
        button.clicked.connect(w.close)

if __name__ == '__main__':
    # setTheme(Theme.DARK)

    app = QApplication(sys.argv)
    w = Window()
    w.show()
    app.exec()
